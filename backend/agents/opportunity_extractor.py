"""
Opportunity Extractor - Extract structured data from "Call for Interest" type emails.

Detects patterns like:
- Call for Interest
- Call for Proposals (CFP)
- Request for Proposals (RFP)
- Call for Applications
- Expression of Interest (EOI)
- Grant Opportunity
- Funding Call

Extracts fields: title, deadline, requirements, contact, budget, eligibility, etc.
Exports to Excel for easy tracking.
"""
from __future__ import annotations
import re
import json
import uuid
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict
import httpx

from ..config import K2_API_KEY, K2_THINK_BASE_URL, K2_THINK_MODEL

# Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[OpportunityExtractor] openpyxl not installed - Excel export disabled")


# Patterns that indicate an opportunity email
OPPORTUNITY_PATTERNS = [
    r"call\s+for\s+interest",
    r"call\s+for\s+proposal",
    r"call\s+for\s+application",
    r"request\s+for\s+proposal",
    r"expression[s]?\s+of\s+interest",
    r"grant\s+opportunity",
    r"funding\s+call",
    r"funding\s+opportunity",
    r"cfp\s*:",
    r"rfp\s*:",
    r"eoi\s*:",
    r"open\s+call",
    r"invitation\s+to\s+apply",
    r"seeking\s+proposals",
    r"submission\s+deadline",
    r"industry\s+collaboration",
    r"partnership\s+opportunity",
    r"research\s+collaboration",
    r"request.*expressions?\s+of\s+interest",
    r"interested\s+in\s+participating",
    r"exploratory\s+meeting",
    r"nominate.*faculty",
    r"respond\s+by",
]


def is_opportunity_email(subject: str, body: str = "") -> bool:
    """Check if an email matches opportunity patterns."""
    text = f"{subject} {body}".lower()
    for pattern in OPPORTUNITY_PATTERNS:
        if re.search(pattern, text, re.IGNORECASE):
            return True
    return False


def get_opportunity_type(subject: str, body: str = "") -> str:
    """Determine the type of opportunity."""
    text = f"{subject} {body}".lower()

    if re.search(r"call\s+for\s+interest|expression\s+of\s+interest|eoi", text):
        return "Call for Interest"
    elif re.search(r"call\s+for\s+proposal|cfp|request\s+for\s+proposal|rfp", text):
        return "Call for Proposals"
    elif re.search(r"call\s+for\s+application|invitation\s+to\s+apply", text):
        return "Call for Applications"
    elif re.search(r"grant|funding", text):
        return "Funding Opportunity"
    else:
        return "Opportunity"


async def extract_opportunity_fields(subject: str, body: str, sender_email: str = "") -> dict:
    """
    Use K2-Think to extract structured fields from an opportunity email.

    Returns:
        {
            "title": str,
            "type": str,
            "deadline": str,
            "organization": str,
            "contact_email": str,
            "contact_name": str,
            "budget": str,
            "eligibility": str,
            "requirements": list[str],
            "key_dates": list[dict],
            "submission_link": str,
            "summary": str,
            "raw_subject": str,
            "raw_body": str,
        }
    """
    opportunity_type = get_opportunity_type(subject, body)

    prompt = f"""You are an expert at extracting structured information from opportunity announcements (calls for interest, proposals, grants, etc.).

Analyze this email and extract all relevant details:

SUBJECT: {subject}

BODY:
{body}

SENDER: {sender_email}

Extract the following fields in JSON format:
{{
    "title": "The official title of the opportunity/call",
    "type": "Type of opportunity (Call for Interest, Call for Proposals, Grant, etc.)",
    "deadline": "Submission deadline in YYYY-MM-DD format if possible, or exact text",
    "organization": "The issuing organization/institution",
    "contact_email": "Contact email address for inquiries",
    "contact_name": "Contact person name",
    "budget": "Funding amount or budget range if mentioned",
    "eligibility": "Who can apply (eligibility criteria)",
    "requirements": ["List", "of", "key", "requirements"],
    "key_dates": [
        {{"event": "Information session", "date": "YYYY-MM-DD or text"}},
        {{"event": "Submission deadline", "date": "YYYY-MM-DD or text"}}
    ],
    "submission_link": "URL for submission portal if provided",
    "summary": "Brief 2-3 sentence summary of the opportunity"
}}

Important:
- Use null for any field you cannot find
- Extract dates precisely when available
- Include ALL requirements mentioned
- For deadline, prioritize the main submission deadline

Return ONLY valid JSON, no additional text."""

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                f"{K2_THINK_BASE_URL}/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {K2_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": K2_THINK_MODEL,
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 2000,
                    "temperature": 0.1,
                },
            )
            response.raise_for_status()
            data = response.json()
            content = data["choices"][0]["message"]["content"]

            # Parse JSON from response
            # Handle potential markdown code blocks
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0]
            elif "```" in content:
                content = content.split("```")[1].split("```")[0]

            extracted = json.loads(content.strip())

            # Add metadata
            extracted["type"] = extracted.get("type") or opportunity_type
            extracted["raw_subject"] = subject
            extracted["raw_body"] = body[:5000]  # Truncate for storage

            return extracted

    except Exception as e:
        print(f"[OpportunityExtractor] AI extraction failed: {e}, using regex fallback")
        # Fallback to regex-based extraction
        return extract_opportunity_with_regex(subject, body, sender_email)


def extract_opportunity_with_regex(subject: str, body: str, sender_email: str = "") -> dict:
    """
    Fallback extraction using regex patterns when AI is unavailable.
    """
    opportunity_type = get_opportunity_type(subject, body)
    full_text = f"{subject}\n{body}"

    # Extract deadline patterns
    deadline = None
    deadline_patterns = [
        r"(?:respond|submit|deadline|due).*?(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})",
        r"(?:respond|submit|deadline|due).*?(\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{2,4})",
        r"\((\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\)",
        r"by\s+(?:early\s+)?(?:next\s+)?(?:week|monday|tuesday|wednesday|thursday|friday)?\s*\(?(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\)?",
    ]
    for pattern in deadline_patterns:
        match = re.search(pattern, full_text, re.IGNORECASE)
        if match:
            deadline = match.group(1)
            break

    # Extract organization
    organization = None
    org_patterns = [
        r"(?:industry\s+partner|organization|company|institution):\s*([^\n]+)",
        r"(?:from|issued\s+by):\s*([^\n]+)",
    ]
    for pattern in org_patterns:
        match = re.search(pattern, full_text, re.IGNORECASE)
        if match:
            organization = match.group(1).strip()
            break

    # Extract email from body
    contact_email = sender_email
    email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', body)
    if email_match:
        contact_email = email_match.group(0)

    # Extract requirements (bullet points)
    requirements = []
    bullet_patterns = [
        r'[-•*]\s*([^\n]+)',
        r'\d+\.\s*([^\n]+)',
    ]
    for pattern in bullet_patterns:
        matches = re.findall(pattern, body)
        requirements.extend([m.strip() for m in matches if len(m.strip()) > 10])

    # Extract URLs
    submission_link = None
    url_match = re.search(r'https?://[^\s<>"{}|\\^\[\]`]+', body)
    if url_match:
        submission_link = url_match.group(0)

    return {
        "title": subject,
        "type": opportunity_type,
        "deadline": deadline,
        "organization": organization,
        "contact_email": contact_email,
        "contact_name": None,
        "budget": None,
        "eligibility": None,
        "requirements": requirements[:10],  # Limit to 10
        "key_dates": [{"event": "Deadline", "date": deadline}] if deadline else [],
        "submission_link": submission_link,
        "summary": body[:500] if body else None,
        "raw_subject": subject,
        "raw_body": body[:5000],
    }


# ──────────────────────────────────────────────────────
# DATABASE OPERATIONS
# ──────────────────────────────────────────────────────

DB_PATH = Path(__file__).parent.parent.parent / "data" / "teamai.db"


def _get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_opportunities_table():
    """Create opportunities table if not exists."""
    conn = _get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS opportunities (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            type TEXT,
            deadline TEXT,
            organization TEXT,
            contact_email TEXT,
            contact_name TEXT,
            budget TEXT,
            eligibility TEXT,
            requirements TEXT,
            key_dates TEXT,
            submission_link TEXT,
            summary TEXT,
            raw_subject TEXT,
            raw_body TEXT,
            department TEXT,
            status TEXT DEFAULT 'new',
            created_at TEXT,
            updated_at TEXT
        )
    """)
    conn.commit()
    conn.close()


def save_opportunity(opportunity: dict, department: str = "general") -> str:
    """Save an extracted opportunity to the database."""
    init_opportunities_table()

    opp_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()

    conn = _get_db()
    conn.execute("""
        INSERT INTO opportunities (
            id, title, type, deadline, organization, contact_email, contact_name,
            budget, eligibility, requirements, key_dates, submission_link, summary,
            raw_subject, raw_body, department, status, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        opp_id,
        opportunity.get("title"),
        opportunity.get("type"),
        opportunity.get("deadline"),
        opportunity.get("organization"),
        opportunity.get("contact_email"),
        opportunity.get("contact_name"),
        opportunity.get("budget"),
        opportunity.get("eligibility"),
        json.dumps(opportunity.get("requirements", [])),
        json.dumps(opportunity.get("key_dates", [])),
        opportunity.get("submission_link"),
        opportunity.get("summary"),
        opportunity.get("raw_subject"),
        opportunity.get("raw_body"),
        department,
        "new",
        now,
        now,
    ))
    conn.commit()
    conn.close()

    return opp_id


def get_all_opportunities(department: Optional[str] = None, status: Optional[str] = None) -> List[dict]:
    """Get all opportunities, optionally filtered by department or status."""
    init_opportunities_table()

    conn = _get_db()
    query = "SELECT * FROM opportunities WHERE 1=1"
    params = []

    if department:
        query += " AND department = ?"
        params.append(department)
    if status:
        query += " AND status = ?"
        params.append(status)

    query += " ORDER BY created_at DESC"

    rows = conn.execute(query, params).fetchall()
    conn.close()

    opportunities = []
    for row in rows:
        opp = dict(row)
        # Parse JSON fields
        opp["requirements"] = json.loads(opp.get("requirements") or "[]")
        opp["key_dates"] = json.loads(opp.get("key_dates") or "[]")
        opportunities.append(opp)

    return opportunities


def update_opportunity_status(opp_id: str, status: str) -> bool:
    """Update opportunity status (new, reviewing, applied, rejected, won)."""
    conn = _get_db()
    result = conn.execute(
        "UPDATE opportunities SET status = ?, updated_at = ? WHERE id = ?",
        (status, datetime.utcnow().isoformat(), opp_id)
    )
    conn.commit()
    conn.close()
    return result.rowcount > 0


# ──────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────

def export_to_excel(opportunities: List[dict], output_path: str = None) -> str:
    """
    Export opportunities to an Excel file.

    Args:
        opportunities: List of opportunity dicts
        output_path: Path to save Excel file (default: data/opportunities.xlsx)

    Returns:
        Path to the created Excel file
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    if output_path is None:
        output_path = str(Path(__file__).parent.parent.parent / "data" / "opportunities.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    # Define columns
    columns = [
        ("Title", 50),
        ("Type", 20),
        ("Deadline", 15),
        ("Organization", 30),
        ("Contact Email", 30),
        ("Contact Name", 20),
        ("Budget", 20),
        ("Eligibility", 40),
        ("Requirements", 50),
        ("Submission Link", 40),
        ("Summary", 60),
        ("Status", 12),
        ("Department", 15),
        ("Created", 20),
    ]

    # Header styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Status colors
    status_colors = {
        "new": "E0E7FF",      # Light blue
        "reviewing": "FEF3C7", # Light yellow
        "applied": "D1FAE5",   # Light green
        "rejected": "FEE2E2",  # Light red
        "won": "A7F3D0",       # Green
    }

    # Write data rows
    for row_idx, opp in enumerate(opportunities, 2):
        # Format requirements as bullet list
        requirements = opp.get("requirements", [])
        if isinstance(requirements, list):
            requirements_str = "\n".join(f"• {r}" for r in requirements)
        else:
            requirements_str = str(requirements)

        row_data = [
            opp.get("title", ""),
            opp.get("type", ""),
            opp.get("deadline", ""),
            opp.get("organization", ""),
            opp.get("contact_email", ""),
            opp.get("contact_name", ""),
            opp.get("budget", ""),
            opp.get("eligibility", ""),
            requirements_str,
            opp.get("submission_link", ""),
            opp.get("summary", ""),
            opp.get("status", "new"),
            opp.get("department", ""),
            opp.get("created_at", "")[:10] if opp.get("created_at") else "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

            # Color code status column
            if col_idx == 12:  # Status column
                status = str(value).lower()
                if status in status_colors:
                    cell.fill = PatternFill(
                        start_color=status_colors[status],
                        end_color=status_colors[status],
                        fill_type="solid"
                    )

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save
    wb.save(output_path)
    print(f"[OpportunityExtractor] Exported {len(opportunities)} opportunities to {output_path}")

    return output_path


# ──────────────────────────────────────────────────────
# HIGH-LEVEL API
# ──────────────────────────────────────────────────────

async def process_opportunity_email(email_data: dict, department: str = "general") -> dict:
    """
    Process an email as an opportunity if it matches patterns.

    Args:
        email_data: Dict with subject, sender, body
        department: Department to associate with

    Returns:
        {
            "is_opportunity": bool,
            "opportunity_id": str or None,
            "extraction": dict or None,
        }
    """
    subject = email_data.get("subject", "")
    body = email_data.get("body", "")
    sender_email = email_data.get("sender", {}).get("email", "")

    if not is_opportunity_email(subject, body):
        return {"is_opportunity": False, "opportunity_id": None, "extraction": None}

    print(f"[OpportunityExtractor] Detected opportunity email: {subject}")

    # Extract fields
    extraction = await extract_opportunity_fields(subject, body, sender_email)

    # Save to database
    opp_id = save_opportunity(extraction, department)

    # Export updated Excel
    try:
        opportunities = get_all_opportunities()
        export_to_excel(opportunities)
    except Exception as e:
        print(f"[OpportunityExtractor] Excel export failed: {e}")

    return {
        "is_opportunity": True,
        "opportunity_id": opp_id,
        "extraction": extraction,
    }
